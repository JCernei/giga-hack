from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from PyPDF2 import PdfReader
import docx 
from PIL import Image
import pytesseract
import openpyxl
import xlrd
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import textwrap
import json
import os
import time
from ai_processing import process_contract
from pdf_generator import generate_invoice_from_text

app = Flask(__name__)
CORS(app)

# Folder to save uploaded files, text files, and generated invoices
UPLOAD_FOLDER = './uploads'
TEXT_FOLDER = './text_files'
INVOICE_FOLDER = './invoices'

# Ensure the folders exist
for folder in [UPLOAD_FOLDER, TEXT_FOLDER, INVOICE_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# Helper function to extract text from a PDF file
def extract_text_from_pdf(file_path):
    reader = PdfReader(file_path)
    text = ''
    for page in reader.pages:
        text += page.extract_text()  # Extract text from each page
    return text

# Helper function to extract text from a Word (.docx) file
def extract_text_from_word(file_path):
    doc = docx.Document(file_path)
    text = ''
    for para in doc.paragraphs:
        text += para.text + '\n'
    return text

# Helper function to extract text from an image (JPG, PNG) file using OCR
def extract_text_from_image(file_path):
    image = Image.open(file_path)
    text = pytesseract.image_to_string(image)
    return text

# Helper function to extract text from an Excel (.xlsx) file
def extract_text_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    text = ''
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += '\t'.join([str(cell) if cell is not None else '' for cell in row]) + '\n'
    return text

# Helper function to extract text from an Excel (.xls) file
def extract_text_from_xls(file_path):
    workbook = xlrd.open_workbook(file_path)
    text = ''
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        for row_idx in range(sheet.nrows):
            row_values = sheet.row_values(row_idx)
            text += '\t'.join([str(cell) for cell in row_values]) + '\n'
    return text

@app.route('/api/convert-contract', methods=['POST'])
def convert_contract():
    if 'contract' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['contract']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    # Save the uploaded file
    file_ext = os.path.splitext(file.filename)[1].lower()
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    # Extract text based on file type
    text = ""
    if file_ext == '.pdf':
        text = extract_text_from_pdf(filepath)
    elif file_ext == '.docx':
        text = extract_text_from_word(filepath)
    elif file_ext in ['.jpg', '.jpeg', '.png']:
        text = extract_text_from_image(filepath)
    elif file_ext == '.xlsx':
        text = extract_text_from_excel(filepath)
    elif file_ext == '.xls':
        text = extract_text_from_xls(filepath)
    else:
        return jsonify({'error': 'Unsupported file type'}), 400

    # Save the extracted text into a .txt file
    text_filename = os.path.splitext(file.filename)[0] + '.txt'
    text_filepath = os.path.join(TEXT_FOLDER, text_filename)
    with open(text_filepath, 'w') as text_file:
        text_file.write(text)

    # Process the extracted text with the AI model
    processed_text = process_contract(text)
    
    # Generate a unique invoice based on the uploaded file
    invoice_path = generate_invoice_from_text(processed_text, os.path.splitext(file.filename)[0])

    # Return the URL to download the generated invoice and the extracted text file
    return jsonify({
        'invoiceUrl': f'/api/download-invoice/{os.path.basename(invoice_path)}'
    })

@app.route('/api/download-invoice/<filename>', methods=['GET'])
def download_invoice(filename):
    # Send the specified invoice file
    invoice_path = os.path.join(INVOICE_FOLDER, filename)
    if os.path.exists(invoice_path):
        return send_file(invoice_path, as_attachment=True, download_name=filename, mimetype='application/pdf')
    else:
        return jsonify({'error': 'File not found'}), 404

@app.route('/api/download-text/<filename>', methods=['GET'])
def download_text(filename):
    # Send the extracted text file
    text_filepath = os.path.join(TEXT_FOLDER, filename)
    if os.path.exists(text_filepath):
        return send_file(text_filepath, as_attachment=True, download_name=filename)
    else:
        return jsonify({'error': 'File not found'}), 404

if __name__ == '__main__':
    app.run(debug=True)
